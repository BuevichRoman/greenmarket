from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from app.parsing.exceptions import ExcelParserError, ParserError
from app.parsing.excel_parser import ExcelParser


def make_xlsx(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    """Строит реальный .xlsx-файл через openpyxl — используется вместо mock."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    return path


def test_parses_valid_workbook_into_raw_workbook(tmp_path):
    path = make_xlsx(
        tmp_path / "valid.xlsx",
        {
            "Каталог": [["SellerProductId", "Цена"], [1, 99.5]],
            "Товарные группы": [["Овощи"]],
        },
    )

    result = ExcelParser().parse(path)

    assert result.source == "valid.xlsx"
    assert [sheet.name for sheet in result.sheets] == ["Каталог", "Товарные группы"]
    assert result.sheets[0].rows == [["SellerProductId", "Цена"], [1, 99.5]]
    assert result.sheets[1].rows == [["Овощи"]]


def test_broken_file_raises_excel_parser_error(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"this is not a real xlsx file")

    with pytest.raises(ExcelParserError):
        ExcelParser().parse(path)


def test_excel_parser_error_is_a_parser_error():
    assert issubclass(ExcelParserError, ParserError)


def test_sheet_index_matches_position_in_workbook(tmp_path):
    path = make_xlsx(
        tmp_path / "multi.xlsx",
        {"First": [[1]], "Second": [[2]], "Third": [[3]]},
    )

    result = ExcelParser().parse(path)

    assert [(sheet.name, sheet.index) for sheet in result.sheets] == [
        ("First", 0),
        ("Second", 1),
        ("Third", 2),
    ]


def test_preserves_blank_row_in_the_middle_of_data(tmp_path):
    path = make_xlsx(
        tmp_path / "with_gap.xlsx",
        {"Каталог": [[1, "A"], [None, None], [3, "C"]]},
    )

    result = ExcelParser().parse(path)

    assert result.sheets[0].rows == [[1, "A"], [None, None], [3, "C"]]


def test_reads_formula_as_raw_string_not_computed_value(tmp_path):
    path = tmp_path / "formulas.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Каталог"
    sheet["A1"] = "=1+1"
    workbook.save(path)

    result = ExcelParser().parse(path)

    assert result.sheets[0].rows == [["=1+1"]]


def test_reads_merged_cells_without_filling_in_values(tmp_path):
    path = tmp_path / "merged.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Каталог"
    sheet["A1"] = "merged value"
    sheet.merge_cells("A1:B1")
    workbook.save(path)

    result = ExcelParser().parse(path)

    assert result.sheets[0].rows == [["merged value", None]]


def test_reads_dates_as_datetime_objects_not_strings(tmp_path):
    path = tmp_path / "dates.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Каталог"
    sheet["A1"] = date(2026, 7, 10)
    workbook.save(path)

    result = ExcelParser().parse(path)

    assert result.sheets[0].rows == [[datetime(2026, 7, 10)]]


def test_parses_empty_sheet_without_error(tmp_path):
    path = make_xlsx(tmp_path / "empty.xlsx", {"Каталог": []})

    result = ExcelParser().parse(path)

    assert result.sheets[0].name == "Каталог"
    assert result.sheets[0].rows == []
