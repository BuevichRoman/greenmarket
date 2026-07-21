from pathlib import Path

import openpyxl
from sqlalchemy import text

from app.infrastructure.repositories.product_group_repository import ProductGroupRepository
from app.infrastructure.repositories.product_repository import ProductRepository
from app.mapping.mapper import Mapper
from app.parsing.excel_parser import ExcelParser
from app.parsing.raw_workbook import RawSheet, RawWorkbook
from app.platform.photo_gateway import PhotoGateway
from app.validation.business_validator import BusinessValidator
from app.validation.semantic_validator import SemanticValidator
from app.validation.structure_validator import CATALOG_COLUMNS, CATALOG_SHEET, SYSTEM_SHEET, StructureValidator
from app.validation.validator import Validator

REPO_ROOT = Path(__file__).resolve().parents[2]
MASTER_TEMPLATE_PATH = REPO_ROOT / "docs" / "02-domain" / "templates" / "catalog_template_v1.xlsx"
EXAMPLES_DIR = REPO_ROOT / "docs" / "02-domain" / "templates" / "examples"
PARTIAL_EXAMPLE_PATH = EXAMPLES_DIR / "catalog_template_v1_partial.xlsx"
FULL_EXAMPLE_PATH = EXAMPLES_DIR / "catalog_template_v1_full.xlsx"


def _make_validator(session) -> Validator:
    return Validator(
        StructureValidator(),
        SemanticValidator(ProductGroupRepository(session), ProductRepository(session), PhotoGateway(session)),
        BusinessValidator(),
    )


_PHOTO_COLUMN_INDEX = len(CATALOG_COLUMNS) - 1


def _with_real_photo_ids(workbook: RawWorkbook, session) -> RawWorkbook:
    """Примеры шаблона (.xlsx) хранят символические id в колонке «Фото»,
    которых нет в свежей БД теста/CI — подставляет реально вставленные Photo
    перед прогоном через SemanticValidator, не трогая остальные листы."""
    catalog = next(sheet for sheet in workbook.sheets if sheet.name == CATALOG_SHEET)
    new_rows = [catalog.rows[0]]
    for row_index, row in enumerate(catalog.rows[1:]):
        photo_id = session.execute(
            text("INSERT INTO Photo (s3_key) VALUES (:s3_key)"), {"s3_key": f"artifact-{workbook.source}-{row_index}.jpg"}
        ).lastrowid
        new_row = list(row)
        if len(new_row) > _PHOTO_COLUMN_INDEX:
            new_row[_PHOTO_COLUMN_INDEX] = str(photo_id)
        new_rows.append(new_row)
    patched_catalog = RawSheet(name=catalog.name, index=catalog.index, rows=new_rows)
    sheets = [patched_catalog if s.name == CATALOG_SHEET else s for s in workbook.sheets]
    return RawWorkbook(source=workbook.source, sheets=sheets)


def test_master_template_file_exists():
    assert MASTER_TEMPLATE_PATH.is_file(), "Запусти `uv run python -m app.catalog_template.build` и закоммить файл"


def test_master_template_passes_structure_validation():
    # "Новый пустой шаблон" из акцептанс-критериев PR-008: 0 строк каталога,
    # структура должна быть валидна как есть, до заполнения продавцом.
    workbook = ExcelParser().parse(MASTER_TEMPLATE_PATH)

    result = StructureValidator().validate(workbook)

    assert result.is_valid, result.errors


def test_master_template_retains_formatting_and_protection():
    # ExcelParser (используется в остальных тестах этого файла) читает только
    # значения ячеек и не видит комментарии/защиту/валидацию — эти свойства
    # уже проверены на build_workbook() в test_catalog_template_builder.py,
    # но не на самом закоммиченном бинарнике. Этот тест — недостающее звено:
    # доказывает, что .xlsx в репозитории на деле несёт те же
    # усability-свойства, а не был пересохранён без них.
    workbook = openpyxl.load_workbook(MASTER_TEMPLATE_PATH)

    catalog_sheet = workbook[CATALOG_SHEET]
    for col_index in range(1, len(CATALOG_COLUMNS) + 1):
        assert catalog_sheet.cell(row=1, column=col_index).comment is not None
    assert catalog_sheet.column_dimensions["A"].hidden is True
    assert catalog_sheet.protection.sheet is False

    system_sheet = workbook[SYSTEM_SHEET]
    assert system_sheet.protection.sheet is True
    assert system_sheet.sheet_state == "hidden"

    hard_enforced_ranges = {"C2", "D2", "E2", "G2"}
    validations = catalog_sheet.data_validations.dataValidation
    assert len(validations) == 5
    for dv in validations:
        sqref = str(dv.sqref)
        if any(sqref.startswith(prefix) for prefix in hard_enforced_ranges):
            assert dv.showErrorMessage is True

    assert catalog_sheet.freeze_panes == "A2"
    assert str(catalog_sheet.auto_filter.ref) == "A1:J1000"
    assert catalog_sheet.column_dimensions["B"].width >= len("Наименование продавца") * 0.9


def test_partial_example_passes_full_pipeline(session):
    workbook = ExcelParser().parse(PARTIAL_EXAMPLE_PATH)
    workbook = _with_real_photo_ids(workbook, session)

    result = _make_validator(session).validate(workbook)
    assert result.is_valid, result.errors

    model = Mapper().map(workbook, result, seller_id=1)
    assert len(model.products) == 2


def test_full_example_passes_full_pipeline(session):
    workbook = ExcelParser().parse(FULL_EXAMPLE_PATH)
    workbook = _with_real_photo_ids(workbook, session)

    result = _make_validator(session).validate(workbook)
    assert result.is_valid, result.errors

    model = Mapper().map(workbook, result, seller_id=1)
    assert len(model.products) == 3
    assert model.products[2].product_name == "Прочее"
