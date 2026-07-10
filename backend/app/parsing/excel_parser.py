from pathlib import Path

import openpyxl
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.worksheet import Worksheet

from app.parsing.exceptions import ExcelParserError
from app.parsing.raw_workbook import RawSheet, RawWorkbook


class ExcelParser:
    """Читает .xlsx в RawWorkbook. Ничего не знает о правилах GreenMarket —
    не проверяет листы/колонки/типы, не интерпретирует и не отбрасывает данные.
    Детерминирован: один и тот же файл всегда даёт один и тот же RawWorkbook.
    """

    def parse(self, path: Path) -> RawWorkbook:
        try:
            return self._parse(path)
        except Exception as exc:
            raise ExcelParserError(f"Не удалось прочитать Excel-файл: {path}") from exc

    def _parse(self, path: Path) -> RawWorkbook:
        workbook = openpyxl.load_workbook(path, data_only=False)
        sheets = [
            RawSheet(name=name, index=index, rows=self._read_rows(workbook[name]))
            for index, name in enumerate(workbook.sheetnames)
        ]
        return RawWorkbook(source=Path(path).name, sheets=sheets)

    def _read_rows(self, sheet: Worksheet | Chartsheet) -> list[list[object]]:
        # Chartsheet — лист-диаграмма (Excel: "переместить диаграмму на
        # отдельный лист"), у него нет ячеек, только сам график.
        if isinstance(sheet, Chartsheet):
            return []
        return [list(row) for row in sheet.iter_rows(values_only=True)]
