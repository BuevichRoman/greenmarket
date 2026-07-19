# backend/app/catalog_template/build.py
"""Строит официальный Catalog Template v1.0 (PR-008).

Единственный источник структуры (листы, колонки, обязательность,
`_System`) — app.validation.structure_validator: этот модуль не хранит
собственной копии контракта, чтобы шаблон не мог разойтись с
Parser/Validator/Mapper.

Порождённый .xlsx — нормативный артефакт (docs/02-domain/Catalog_Template.md).
Google Sheets-копия для продавцов — производная: создаётся импортом этого
файла в Google Drive и может быть пересоздана из него в любой момент; сама
она источником истины не является (решение колонки от 2026-07-13).
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.catalog_template.data import (
    COLUMN_HINTS,
    COLUMN_WIDTHS,
    GROUP_NAMES_FOR_DROPDOWN,
    GROUP_ROWS,
    INSTRUCTION_LINES,
    PRODUCT_NAMES_FOR_DROPDOWN,
    PRODUCT_ROWS,
    TEMPLATE_ID,
    TEMPLATE_VERSION,
)
from app.validation.structure_validator import (
    CATALOG_COLUMNS,
    CATALOG_SHEET,
    INSTRUCTION_SHEET,
    PRODUCT_GROUPS_COLUMNS,
    PRODUCTS_COLUMNS,
    PRODUCTS_SHEET,
    PRODUCT_GROUPS_SHEET,
    SYSTEM_SHEET,
)

OUTPUT_PATH = Path(__file__).resolve().parents[3] / "docs" / "02-domain" / "templates" / "catalog_template_v1.xlsx"

_MAX_DATA_ROW = 1000  # запас строк под форматирование/валидацию для будущих товаров продавца

_REQUIRED_FILL = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
_OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_SERVICE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def build_workbook(
    catalog_rows: list[list[object]] | None = None,
    *,
    group_rows: list[list[object]] | None = None,
    product_rows: list[list[object]] | None = None,
    group_names: list[str] | None = None,
    product_names: list[str] | None = None,
) -> Workbook:
    """group_rows/product_rows/group_names/product_names по умолчанию — статические
    данные data.py (не меняет поведение существующих вызовов). Передать
    live-данные из db_source.py, чтобы собрать шаблон из актуальной БД —
    см. main()/--from-db.
    """
    group_rows = group_rows if group_rows is not None else GROUP_ROWS
    product_rows = product_rows if product_rows is not None else PRODUCT_ROWS
    group_names = group_names if group_names is not None else GROUP_NAMES_FOR_DROPDOWN
    product_names = product_names if product_names is not None else PRODUCT_NAMES_FOR_DROPDOWN

    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_catalog_sheet(workbook, catalog_rows or [], group_names, product_names)
    _build_reference_sheet(workbook, PRODUCT_GROUPS_SHEET, PRODUCT_GROUPS_COLUMNS, group_rows)
    _build_reference_sheet(workbook, PRODUCTS_SHEET, PRODUCTS_COLUMNS, product_rows)
    _build_instruction_sheet(workbook)
    _build_system_sheet(workbook)

    return workbook


def _build_catalog_sheet(
    workbook: Workbook,
    catalog_rows: list[list[object]],
    group_names: list[str],
    product_names: list[str],
) -> None:
    ws = workbook.create_sheet(CATALOG_SHEET, 0)

    for col_index, column in enumerate(CATALOG_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=column.name)
        cell.comment = Comment(COLUMN_HINTS[column.name], "GreenMarket")
        if column.name == "SellerProductId":
            cell.fill = _SERVICE_FILL
        elif column.required:
            cell.fill = _REQUIRED_FILL
        else:
            cell.fill = _OPTIONAL_FILL
        ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS[column.name]

    for row_index, row in enumerate(catalog_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    _apply_catalog_validation(ws, group_names, product_names)
    ws.column_dimensions["A"].hidden = True
    ws.freeze_panes = "A2"
    last_row = max(ws.max_row, _MAX_DATA_ROW)
    last_column_letter = ws.cell(row=1, column=len(CATALOG_COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_column_letter}{last_row}"


def _quoted_list(values: list[str]) -> str:
    return '"' + ",".join(values) + '"'


def _apply_catalog_validation(ws: Worksheet, group_names: list[str], product_names: list[str]) -> None:
    last_row = max(ws.max_row, _MAX_DATA_ROW)

    group_validation = DataValidation(
        type="list", formula1=_quoted_list(group_names), allow_blank=True, showErrorMessage=True
    )
    group_validation.errorTitle = "Неизвестная товарная группа"
    group_validation.error = "Выберите товарную группу из списка (см. лист «Товарные группы»)"
    ws.add_data_validation(group_validation)
    group_validation.add(f"C2:C{last_row}")

    product_validation = DataValidation(
        type="list", formula1=_quoted_list(product_names), allow_blank=True, showErrorMessage=True
    )
    product_validation.errorTitle = "Неизвестная товарная позиция"
    product_validation.error = "Выберите товарную позицию из списка или «Прочее» (см. лист «Товарные позиции»)"
    ws.add_data_validation(product_validation)
    product_validation.add(f"D2:D{last_row}")

    price_validation = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=True
    )
    price_validation.errorTitle = "Некорректная цена"
    price_validation.error = "Цена не может быть отрицательной"
    ws.add_data_validation(price_validation)
    price_validation.add(f"E2:E{last_row}")

    stock_validation = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=True
    )
    stock_validation.errorTitle = "Некорректный остаток"
    stock_validation.error = "Остаток не может быть отрицательным"
    ws.add_data_validation(stock_validation)
    stock_validation.add(f"G2:G{last_row}")

    unit_validation = DataValidation(
        type="list", formula1='"кг,г,шт,уп,л,мл"', allow_blank=True, showErrorMessage=True
    )
    unit_validation.errorStyle = "warning"
    unit_validation.errorTitle = "Необычная единица продажи"
    unit_validation.error = "Обычно используется: кг, г, шт, уп, л, мл. При необходимости можно ввести своё значение."
    ws.add_data_validation(unit_validation)
    unit_validation.add(f"F2:F{last_row}")


def _build_reference_sheet(workbook: Workbook, sheet_name: str, columns, rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(sheet_name)
    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=column.name)
        cell.fill = _SERVICE_FILL
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
    ws.protection.sheet = True


def _build_instruction_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet(INSTRUCTION_SHEET)
    for row_index, line in enumerate(INSTRUCTION_LINES, start=1):
        ws.cell(row=row_index, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    ws.protection.sheet = True


def _build_system_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet(SYSTEM_SHEET)
    ws.append(["TemplateVersion", TEMPLATE_VERSION])
    ws.append(["TemplateId", TEMPLATE_ID])
    ws.protection.sheet = True
    ws.sheet_state = "hidden"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--from-db",
        action="store_true",
        help=(
            "Собрать листы «Товарные группы»/«Товарные позиции» из актуальной БД "
            "(ProductGroup/Product), а не из статического списка в data.py. "
            "Требует доступную БД (см. app.infrastructure.database)."
        ),
    )
    args = parser.parse_args()

    if args.from_db:
        from app.catalog_template.db_source import load_dropdown_names, load_group_rows, load_product_rows
        from app.infrastructure.database import SessionLocal

        session = SessionLocal()
        try:
            group_rows = load_group_rows(session)
            product_rows = load_product_rows(session)
            group_names, product_names = load_dropdown_names(group_rows, product_rows)
        finally:
            session.close()
        workbook = build_workbook(
            group_rows=group_rows, product_rows=product_rows, group_names=group_names, product_names=product_names
        )
    else:
        workbook = build_workbook()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Шаблон сохранён: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
